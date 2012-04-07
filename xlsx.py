from openpyxl.reader.excel import load_workbook
from openpyxl.shared.exc import InvalidFileException
import panscan
import logger
import StringIO

def plugin(fname, data = None):
    try:
        if (data != None):
            flo = StringIO.StringIO(data)
            wb = load_workbook(flo)
        else:
            wb = load_workbook(filename = fname, use_iterators = True)
    except InvalidFileException:
        logger.Logger().log_ignore(fname, "Invalid Excel file")
        return
        
    sheetNames = wb.get_sheet_names()

    for name in sheetNames:
        sheet = wb.get_sheet_by_name(name)
        for cell in sheet.get_cell_collection():
            if (cell.value !=None):
                # scan text for pans and return as a list
                value = str(cell.value)
                
                pans = panscan.panscan(" " + value + " ")

                # log each pan found - bit of a hack to see if the pans list is empty
                for p in pans:
                    logger.Logger().log_pan(fname, p[1], name +" Cell " + cell.column +":"+ str(cell.row))
    
def intToChar(integer):
    """A really lazy way to convert the ints to chars. will return the column name according to the int provided"""
    temp = integer/26 +1
    character = integer%26
    character+=65    
    return chr(character)*temp


#plugin('/home/d/Dropbox/Uni/3000/ccsrch/testFiles/2 Excel.xlsx')
